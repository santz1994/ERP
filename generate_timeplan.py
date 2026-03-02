"""
Generate ERP Project Timeplan Excel
PT Quty Karunia ERP 2026-2027
Solo Developer - Target: 1-2 years
Current progress: Session 51+ (as of March 2, 2026)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import date, timedelta
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def hdr(text, r, c, ws, bold=True, size=11, bg=None, fg="000000", align="center", wrap=False, colspan=None):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if colspan:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + colspan - 1)
    return cell

def val(v, r, c, ws, align="center", bold=False, size=10, bg=None, fg="000000", wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

COLORS = {
    "done":       "4CAF50",
    "in_prog":    "2196F3",
    "todo":       "9E9E9E",
    "buyer":      "AB47BC",
    "warn":       "FF9800",
    "header":     "1565C0",
    "subheader":  "283593",
    "phase1":     "1B5E20",
    "phase2":     "0D47A1",
    "phase3":     "4A148C",
    "phase4":     "BF360C",
    "phase5":     "006064",
    "weekend":    "EEEEEE",
    "today":      "FFEB3B",
    "row_even":   "F8F9FA",
    "row_odd":    "FFFFFF",
}

def freeze_and_style(ws, row=2, col=2):
    ws.freeze_panes = ws.cell(row=row, column=col)
    ws.sheet_view.zoomScale = 90

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: PROJECT OVERVIEW / GANTT
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "📊 Project Gantt"
ws.sheet_properties.tabColor = "1565C0"

# Title
ws.merge_cells("A1:Z1")
c = ws["A1"]
c.value = "ERP PT Quty Karunia 2026-2027 — Project Gantt & Roadmap"
c.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor=COLORS["header"])
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Column headers
headers = ["Phase", "Module / Feature", "Est. Sessions", "Start", "End", "Status", "Progress %", "Notes"]
col_widths = [8, 38, 14, 12, 12, 14, 14, 40]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    hdr(h, 2, i, ws, bg=COLORS["subheader"], fg="FFFFFF", size=10)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[2].height = 22

# Project phases and features
phases = [
    # (phase_num, display, module_feature, est_sessions, start, end, status, pct, notes)
    # ── PHASE 1: FOUNDATION (DONE) ──────────────────────────────────────────
    ("1", "PHASE 1 — FOUNDATION", "", "", "", "", "✅ DONE", 100, "Completed ~Session 30"),
    ("1", "Authentication & RBAC", "Login, JWT, Role & Permission system", 8, "2026-01-01", "2026-01-31", "✅ Done", 100, "All 11 roles, fine-grained permissions"),
    ("1", "Database Schema", "PostgreSQL, SQLAlchemy ORM, Base models", 5, "2026-01-01", "2026-01-20", "✅ Done", 100, "Auto-migrate on startup"),
    ("1", "Purchasing Module", "PO KAIN / LABEL / ACC, Dual Trigger, multi-batch", 10, "2026-01-15", "2026-02-10", "✅ Done", 100, "2 MO auto-created per PO KAIN"),
    ("1", "Warehouse", "Material receipt, stock, FG, stock opname", 8, "2026-01-20", "2026-02-15", "✅ Done", 100, "FIFO, lot tracking"),
    ("1", "BOM (Dual)", "BOM Production + BOM Purchasing, import from Excel", 6, "2026-01-25", "2026-02-05", "✅ Done", 100, "5845 BOM lines imported"),
    # ── PHASE 2: CORE MODULES ───────────────────────────────────────────────
    ("2", "PHASE 2 — CORE MODULES", "", "", "", "", "🔵 In Progress", 45, "Current focus"),
    ("2", "PPIC — MO & WO", "MO BUYER/PRODUCTION, WO per dept, SPK auto", 10, "2026-02-01", "2026-03-15", "🔵 In Progress", 65, "Dual MO, WO auto-creation on PO send"),
    ("2", "Production Floor", "Cutting/Emb/Sewing/Finishing/Packing input pages", 12, "2026-02-10", "2026-03-30", "🔵 In Progress", 40, "Daily Input per dept, permission-gated"),
    ("2", "BOM Management UI", "Admin CRUD page for BOM headers & lines", 3, "2026-02-20", "2026-03-05", "✅ Done", 100, "Session 51+"),
    ("2", "Masterdata CRUD", "Products, Categories, Suppliers admin pages", 3, "2026-02-20", "2026-03-05", "✅ Done", 100, "Session 51+"),
    ("2", "PO Delete/Approval", "Request delete workflow, approval by atasan", 2, "2026-02-25", "2026-03-02", "✅ Done", 100, "Session 51+"),
    ("2", "Quality Control", "QC Inspection, Lab Test, Rework workflow", 8, "2026-03-01", "2026-03-31", "⏳ Todo", 5, "QC checkpoint, defect categories"),
    # ── PHASE 3: OPERATIONS ─────────────────────────────────────────────────
    ("3", "PHASE 3 — OPERATIONS", "", "", "", "", "⏳ Todo", 0, "Est. April–June 2026"),
    ("3", "Material Debt System", "Negative inventory, settlement, auto-track", 6, "2026-04-01", "2026-04-20", "⏳ Todo", 0, ""),
    ("3", "WIP Tracking", "WIP Dashboard, inter-dept transfer, buffer", 8, "2026-04-10", "2026-05-10", "⏳ Todo", 0, ""),
    ("3", "Kanban Board", "Real-time Kanban for production floor", 5, "2026-05-01", "2026-05-20", "⏳ Todo", 0, ""),
    ("3", "Finish Goods", "FG stock, dispatch, customer delivery", 6, "2026-05-15", "2026-06-10", "⏳ Todo", 0, ""),
    ("3", "Material Allocation (SPK)", "Auto-allocate materials from WH to SPK", 5, "2026-05-20", "2026-06-15", "⏳ Todo", 0, ""),
    # ── PHASE 4: REPORTING & ANALYTICS ──────────────────────────────────────
    ("4", "PHASE 4 — REPORTING & ANALYTICS", "", "", "", "", "⏳ Todo", 0, "Est. July–Sept 2026"),
    ("4", "Dashboard Analytics", "KPI cards, real-time charts, trend analysis", 8, "2026-07-01", "2026-07-31", "⏳ Todo", 0, ""),
    ("4", "Production Reports", "Daily output, efficiency, rejection rate", 6, "2026-07-15", "2026-08-15", "⏳ Todo", 0, ""),
    ("4", "Purchasing Reports", "PO status, supplier performance, lead time", 4, "2026-08-01", "2026-08-20", "⏳ Todo", 0, ""),
    ("4", "QC Reports", "Defect trends, batch pass/fail rates, IKEA compliance", 4, "2026-08-15", "2026-09-05", "⏳ Todo", 0, ""),
    ("4", "Audit Trail", "Full user activity log, change history", 3, "2026-09-01", "2026-09-15", "⏳ Todo", 0, ""),
    # ── PHASE 5: FINISH & DEPLOY ─────────────────────────────────────────────
    ("5", "PHASE 5 — HARDENING & DEPLOY", "", "", "", "", "⏳ Todo", 0, "Est. Oct 2026–Jan 2027"),
    ("5", "Performance Optimization", "DB indexes, query optimization, caching", 5, "2026-10-01", "2026-10-20", "⏳ Todo", 0, ""),
    ("5", "E2E Testing", "Playwright tests, API tests, integration", 6, "2026-10-15", "2026-11-10", "⏳ Todo", 0, ""),
    ("5", "Docker + CI/CD", "Production deploy, staging env, nginx SSL", 4, "2026-11-01", "2026-11-20", "⏳ Todo", 0, ""),
    ("5", "User Training", "SOP documentation, training materials", 5, "2026-11-15", "2026-12-10", "⏳ Todo", 0, ""),
    ("5", "Live Launch", "Full production go-live at PT Quty Karunia", 3, "2026-12-15", "2027-01-15", "⏳ Todo", 0, "🎯 Target: Jan 2027"),
]

phase_colors = {
    "1": ("E8F5E9", COLORS["phase1"]),
    "2": ("E3F2FD", COLORS["phase2"]),
    "3": ("F3E5F5", COLORS["phase3"]),
    "4": ("FBE9E7", COLORS["phase4"]),
    "5": ("E0F7FA", COLORS["phase5"]),
}

r = 3
for ph, feature, description, est_sess, po_start, po_end, status, pct, notes in phases:
    bg_light, bg_dark = phase_colors[ph]
    is_phase_header = feature.startswith("PHASE")

    if is_phase_header:
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=bg_dark)
        val(f"Phase {ph}", r, 1, ws, bold=True, fg="FFFFFF", bg=bg_dark)
        val(feature, r, 2, ws, align="left", bold=True, fg="FFFFFF", bg=bg_dark, size=11)
        val("", r, 3, ws, bg=bg_dark)
        val("", r, 4, ws, bg=bg_dark)
        val("", r, 5, ws, bg=bg_dark)
        val(status, r, 6, ws, bold=True, fg="FFFFFF", bg=bg_dark)
        # Progress bar cell
        pct_cell = val(f"{pct}%", r, 7, ws, bold=True, fg="FFFFFF", bg=bg_dark)
        val(description or notes, r, 8, ws, align="left", fg="FFFFFF", bg=bg_dark, wrap=True)
        ws.row_dimensions[r].height = 24
    else:
        row_bg = COLORS["row_even"] if r % 2 == 0 else COLORS["row_odd"]
        val(f"  {ph}", r, 1, ws, bg=row_bg)
        val(feature, r, 2, ws, align="left", bg=row_bg, wrap=True)
        val(description if isinstance(description, str) else f"{est_sess} sessions", r, 3, ws, bg=row_bg, wrap=True)
        val(po_start, r, 4, ws, bg=row_bg)
        val(po_end, r, 5, ws, bg=row_bg)

        # Status color
        if "Done" in status or "DONE" in status:
            sc = "FFFFFF"; sb = COLORS["done"]
        elif "Progress" in status:
            sc = "FFFFFF"; sb = COLORS["in_prog"]
        else:
            sc = "555555"; sb = "E0E0E0"
        val(status, r, 6, ws, bold=True, fg=sc, bg=sb)

        # Progress %
        if pct == 100:
            pb = COLORS["done"]
        elif pct > 0:
            pb = COLORS["in_prog"]
        else:
            pb = "E0E0E0"
        val(f"{pct}%", r, 7, ws, bold=pct > 0, bg=pb, fg="FFFFFF" if pct > 0 else "888888")
        val(notes, r, 8, ws, align="left", bg=row_bg, wrap=True)
        ws.row_dimensions[r].height = 20
    r += 1

freeze_and_style(ws, 3, 2)

# Summary at top right
summary_start_row = 3
summary_data = [
    ("📅 Project Start", "January 2026"),
    ("📍 Current Date", "March 2, 2026"),
    ("🔢 Sessions Done", "51+"),
    ("🎯 Target Launch", "January 2027"),
    ("👨‍💻 Developer", "Solo (Daniel Rizaldy)"),
    ("🏭 Client", "PT Quty Karunia"),
]
for i, (k, v_) in enumerate(summary_data):
    hdr(k, summary_start_row + i, 10, ws, bg="283593", fg="FFFFFF", size=9, align="left")
    val(v_, summary_start_row + i, 11, ws, align="left", bg="E8EAF6", size=9)
for c_ in [10, 11]:
    ws.column_dimensions[get_column_letter(c_)].width = 22

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: SESSION LOG / DAILY ACTIVITY
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("📝 Session Log")
ws2.sheet_properties.tabColor = "4CAF50"

ws2.merge_cells("A1:J1")
c = ws2["A1"]
c.value = "Session Activity Log — ERP PT Quty Karunia"
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor=COLORS["phase1"])
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

log_headers = ["Session #", "Date", "Duration (h)", "Phase", "Module", "What was done", "Files Changed", "Status", "Blockers", "Next Steps"]
log_widths  = [10, 12, 12, 8, 18, 45, 35, 12, 28, 35]
for i, (h, w) in enumerate(zip(log_headers, log_widths), 1):
    hdr(h, 2, i, ws2, bg=COLORS["subheader"], fg="FFFFFF", size=10)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[2].height = 22

# Sample past sessions
past_sessions = [
    (1,  "2026-01-05", 4, "1", "Auth",        "Setup project, FastAPI base, JWT auth, User model",               "main.py, users.py, auth.py",             "✅ Done", "", "RBAC"),
    (10, "2026-01-15", 5, "1", "Purchasing",   "PO KAIN/LABEL types, Dual Trigger TRIGGER 1",                    "purchasing.py, manufacturing.py",         "✅ Done", "", "TRIGGER 2"),
    (20, "2026-01-25", 6, "1", "BOM",          "Import dual BOM from Excel, 5845 BOM lines",                     "import_bom_production.py, bom.py",        "✅ Done", "", "BOM Purchasing"),
    (30, "2026-02-05", 5, "1", "Warehouse",    "Stock opname, WH routes, material receipt",                      "warehouse_endpoints.py",                  "✅ Done", "", "Material debt"),
    (40, "2026-02-15", 5, "2", "PPIC",         "MO list, SPK list, material allocation",                         "ppic.py, MOListPage.tsx",                 "✅ Done", "", "WO creation"),
    (50, "2026-02-25", 6, "2", "Purchasing",   "PO Detail modal, multi-batch WO, qty overflow",                  "PODetailModal.tsx, purchasing.py",         "✅ Done", "", "PO delete"),
    (51, "2026-03-01", 7, "2", "Multi-feat",   "PO delete/approval, BOM Mgmt page, Masterdata CRUD, 2 MO types", "ppic.py, schemas.py, MOListPage.tsx, ...", "✅ Done", "", "Daily Input per dept"),
]
for i, row_data in enumerate(past_sessions, 3):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    for j, v_ in enumerate(row_data, 1):
        if j == 8:
            cell = val(v_, i, j, ws2, bg=COLORS["done"] if "Done" in str(v_) else COLORS["in_prog"], fg="FFFFFF")
        else:
            val(v_, i, j, ws2, align="left" if j > 4 else "center", bg=bg, wrap=(j in [6, 7, 9, 10]))
    ws2.row_dimensions[i].height = 32

# Template rows (blank) for future sessions
template_start = 3 + len(past_sessions)
for i in range(template_start, template_start + 150):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    for j in range(1, 11):
        val("", i, j, ws2, bg=bg)
    ws2.row_dimensions[i].height = 20

freeze_and_style(ws2, 3, 2)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: BUG & ISSUE TRACKER
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("🐛 Bug Tracker")
ws3.sheet_properties.tabColor = "F44336"

ws3.merge_cells("A1:K1")
c = ws3["A1"]
c.value = "Bug & Issue Tracker — ERP PT Quty Karunia"
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor="C62828")
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

bug_headers = ["ID", "Found Date", "Priority", "Type", "Module", "Description", "Root Cause", "Status", "Fixed Date", "Session #", "Fix Notes"]
bug_widths  = [8, 12, 10, 12, 16, 42, 35, 12, 12, 10, 35]
for i, (h, w) in enumerate(zip(bug_headers, bug_widths), 1):
    hdr(h, 2, i, ws3, bg="C62828", fg="FFFFFF", size=10)
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[2].height = 22

known_bugs = [
    ("BUG-001", "2026-02-20", "🔴 High",   "Bug",    "Masterdata", "AdminMasterdataPage duplicate code after file edit", "PowerShell truncation issue", "✅ Fixed", "2026-02-20", 51, "Truncated to line 477"),
    ("BUG-002", "2026-02-20", "🟡 Med",    "Bug",    "BOM Mgmt",   "Unused imports causing linter errors in bom_management.py", "Copy-paste leftovers", "✅ Fixed", "2026-02-20", 51, "Removed Category, ProductType"),
    ("BUG-003", "2026-03-01", "🔴 High",   "Feature","PPIC",       "Single MO instead of 2 (BUYER + PRODUCTION)",          "Old architecture single-MO", "✅ Fixed", "2026-03-02", 51, "Dual MO trigger implemented"),
    ("BUG-004", "2026-03-01", "🟡 Med",    "Bug",    "Navigation", "Daily Input generic /daily-production not per dept",    "Old single-page design", "✅ Fixed", "2026-03-02", 51, "Per-dept routes + sidebar"),
    ("BUG-005", "2026-03-01", "🔴 High",   "Bug",    "App.tsx",    "Missing routes for 4 dept input pages (Emb/Sew/Fin/Pack)", "Never registered", "✅ Fixed", "2026-03-02", 51, "4 routes added to App.tsx"),
    ("BUG-006", "2026-03-02", "🟠 Med",    "Verify", "Purchasing", "Buttons in PODetailModal may not update status correctly", "To be tested", "⏳ Open", "", "", "Test Send/Receive/Cancel"),
    ("BUG-007", "2026-03-02", "🟠 Med",    "Verify", "PPIC",       "MOListPage GET api.ppic.getMOs returns shapes correctly?", "API shape unknown", "⏳ Open", "", "", "Verify response mapping"),
]

for i, row_data in enumerate(known_bugs, 3):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    for j, v_ in enumerate(row_data, 1):
        if j == 8:  # Status
            if "Fixed" in str(v_):
                cell = val(v_, i, j, ws3, bg=COLORS["done"], fg="FFFFFF", bold=True)
            elif "Open" in str(v_):
                cell = val(v_, i, j, ws3, bg=COLORS["warn"], fg="FFFFFF", bold=True)
            else:
                val(v_, i, j, ws3, bg=bg)
        elif j == 3:  # Priority
            if "High" in str(v_):
                val(v_, i, j, ws3, bg="FFCDD2", bold=True)
            elif "Med" in str(v_) or "Orange" in str(v_):
                val(v_, i, j, ws3, bg="FFE0B2", bold=True)
            else:
                val(v_, i, j, ws3, bg=bg)
        else:
            val(v_, i, j, ws3, align="left" if j > 4 else "center", bg=bg, wrap=(j in [6, 7, 11]))
    ws3.row_dimensions[i].height = 30

# Blank rows for future bugs
template_start = 3 + len(known_bugs)
for i in range(template_start, template_start + 100):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    for j in range(1, 12):
        val("", i, j, ws3, bg=bg)
    ws3.row_dimensions[i].height = 18

freeze_and_style(ws3, 3, 2)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4: SESSION PROGRESS TRACKER
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("🎯 Progress Tracker")
ws4.sheet_properties.tabColor = "9C27B0"

ws4.merge_cells("A1:H1")
c = ws4["A1"]
c.value = "Session Progress Tracker — ERP PT Quty Karunia (Target: 100–150 Sessions)"
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor="6A1B9A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

prog_headers = ["Session Range", "Phase", "Period", "Key Deliverables", "Sessions Est.", "Sessions Done", "% Done", "Status"]
prog_widths  = [16, 8, 22, 55, 14, 14, 10, 14]
for i, (h, w) in enumerate(zip(prog_headers, prog_widths), 1):
    hdr(h, 2, i, ws4, bg="6A1B9A", fg="FFFFFF", size=10)
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.row_dimensions[2].height = 22

progress_rows = [
    ("1–30",    "1", "Jan 2026",        "Auth, Purchasing, BOM, Warehouse, PPIC basics",             30, 30, 100, "✅ Complete"),
    ("31–50",   "2", "Feb 2026",        "Production Form, QC, Rework, Material Debt, Reports draft", 20, 21, 100, "✅ Complete"),
    ("51–70",   "2", "Mar 2026",        "2 MO types, Daily Input per dept, Fix broken pages/modals, WIP dashboard", 20, 1, 5, "🔵 In Progress"),
    ("71–90",   "3", "Apr–May 2026",    "Material Debt System, WIP Tracking, Kanban, Finish Goods",  20, 0, 0, "⏳ Todo"),
    ("91–110",  "4", "Jun–Jul 2026",    "Analytics Dashboard, Production/Purchasing/QC Reports",     20, 0, 0, "⏳ Todo"),
    ("111–130", "5", "Aug–Oct 2026",    "Performance tuning, E2E Testing, CI/CD, Docs",              20, 0, 0, "⏳ Todo"),
    ("131–150", "5", "Nov 2026–Jan 2027","User training, pilot run, live launch",                     20, 0, 0, "⏳ Todo"),
]

for i, (sess_range, phase, period, deliverables, est, done, pct, status) in enumerate(progress_rows, 3):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    val(sess_range, i, 1, ws4, bold=True, bg=bg)
    val(phase, i, 2, ws4, bg=bg)
    val(period, i, 3, ws4, bg=bg)
    val(deliverables, i, 4, ws4, align="left", bg=bg, wrap=True)
    val(est, i, 5, ws4, bg=bg)
    val(done, i, 6, ws4, bold=True, bg=bg)
    if pct == 100:
        pct_cell = val(f"{pct}%", i, 7, ws4, bold=True, bg=COLORS["done"], fg="FFFFFF")
    elif pct > 0:
        pct_cell = val(f"{pct}%", i, 7, ws4, bold=True, bg=COLORS["in_prog"], fg="FFFFFF")
    else:
        pct_cell = val(f"{pct}%", i, 7, ws4, bg="E0E0E0", fg="888888")
    if "Complete" in status:
        val(status, i, 8, ws4, bold=True, bg=COLORS["done"], fg="FFFFFF")
    elif "Progress" in status:
        val(status, i, 8, ws4, bold=True, bg=COLORS["in_prog"], fg="FFFFFF")
    else:
        val(status, i, 8, ws4, bg="E0E0E0", fg="888888")
    ws4.row_dimensions[i].height = 36

# Overall summary
r_sum = 3 + len(progress_rows) + 2
ws4.merge_cells(start_row=r_sum, start_column=1, end_row=r_sum, end_column=8)
c = ws4.cell(row=r_sum, column=1)
c.value = "📊 OVERALL PROGRESS SUMMARY"
c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor="283593")
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[r_sum].height = 24

summary_kpi = [
    ("Total Sessions Planned", "~150"),
    ("Sessions Completed", "51"),
    ("Overall Progress", "34%"),
    ("Estimated Completion", "January 2027"),
    ("Avg Sessions/Week", "~5"),
    ("Critical Path", "Daily Input + WIP + Reports"),
]
for i, (k, v_) in enumerate(summary_kpi, 1):
    hdr(k, r_sum + i, 1, ws4, bg="3949AB", fg="FFFFFF", align="left", colspan=2)
    val(v_, r_sum + i, 3, ws4, bold=True, bg="E8EAF6", size=11)
    ws4.row_dimensions[r_sum + i].height = 22

freeze_and_style(ws4, 3, 1)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5: WEEKLY PLANNER
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("📅 Weekly Planner")
ws5.sheet_properties.tabColor = "FF9800"

ws5.merge_cells("A1:G1")
c = ws5["A1"]
c.value = "Weekly Development Planner — March 2026 onwards"
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = PatternFill("solid", fgColor="E65100")
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

week_headers = ["Week", "Date Range", "Focus Area", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Weekend/Buffer", "Completed?"]
week_widths  = [8, 20, 24, 28, 28, 28, 28, 28, 24, 12]
for i, (h, w) in enumerate(zip(week_headers, week_widths), 1):
    hdr(h, 2, i, ws5, bg="E65100", fg="FFFFFF", size=9)
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.row_dimensions[2].height = 22

# Starting week: March 2, 2026
start = date(2026, 3, 2)

weekly_plan = [
    ("Fix broken pages/modals, Daily Input pages", 
     "Test & fix PO modal buttons", "Test MO list shows 2 MOs", "Fix broken sidebar routes", "WIP Dashboard", "Sewing/Finishing input pages"),
    ("Production Daily Input — real data flow",
     "Cutting Input API connect", "Embroidery Daily Input", "Sewing Daily Input", "Finishing Daily Input", "Packing Daily Input"),
    ("QC Module",
     "QC Inspection form", "Defect categories CRUD", "Rework workflow UI", "QC checkpoint page", "QC reports"),
    ("Material Debt + WIP",
     "Material debt dashboard", "Settlement workflow", "WIP transfer logic", "WIP dashboard", "Integration test"),
    ("Performance + Polish",
     "DB query optimization", "Pagination on all lists", "Loading states fix", "Form validation polish", "Mobile responsiveness"),
    ("Kanban + FG",
     "Kanban board backend", "Kanban frontend", "FG receipt form", "FG stock display", "FG dispatch"),
    ("Reports Phase 1",
     "Production daily report", "Purchasing summary", "QC defect trends", "Report export PDF", "Dashboard charts"),
    ("Reports Phase 2 + Analytics",
     "KPI dashboard", "IKEA compliance report", "Supplier performance", "Week-over-week trends", "Export to Excel"),
]

r = 3
for wk_i, (focus, mon, tue, wed, thu, fri) in enumerate(weekly_plan):
    week_start = start + timedelta(weeks=wk_i)
    week_end   = week_start + timedelta(days=4)
    bg = COLORS["row_even"] if wk_i % 2 == 0 else COLORS["row_odd"]
    val(f"W{wk_i+1}", r, 1, ws5, bold=True, bg=bg)
    val(f"{week_start.strftime('%b %d')} – {week_end.strftime('%b %d, %Y')}", r, 2, ws5, bg=bg)
    val(focus, r, 3, ws5, align="left", bg=bg, bold=True, wrap=True)
    val(mon, r, 4, ws5, align="left", bg=bg, wrap=True)
    val(tue, r, 5, ws5, align="left", bg=bg, wrap=True)
    val(wed, r, 6, ws5, align="left", bg=bg, wrap=True)
    val(thu, r, 7, ws5, align="left", bg=bg, wrap=True)
    val(fri, r, 8, ws5, align="left", bg=bg, wrap=True)
    val("Review + commit", r, 9, ws5, align="left", bg=bg, fg="777777")
    val("", r, 10, ws5, bg=bg)
    ws5.row_dimensions[r].height = 42
    r += 1

# Blank template weeks
for i in range(r, r + 40):
    bg = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
    for j in range(1, 11):
        val("", i, j, ws5, bg=bg)
    ws5.row_dimensions[i].height = 32

freeze_and_style(ws5, 3, 2)

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
out_path = r"d:\Project\ERP2026\ERP_Project_Timeplan_2026-2027.xlsx"
wb.save(out_path)
print(f"✅ Excel timeplan saved: {out_path}")
print(f"   Tabs: {[ws.title for ws in wb.worksheets]}")
