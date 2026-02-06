"""Masterdata Import Service - Session 49 Phase 8
Handles bulk import of masterdata from Excel templates with validation and transaction safety.

Import Types:
1. Suppliers (partners with type=SUPPLIER)
2. Materials (products with type=RAW_MATERIAL/BAHAN_PENOLONG)
3. Articles (products with type=FINISHED_GOODS)
4. BOM (bill of materials - header + details)
5. Supplier-Material Relations (future: pricing, lead time)

Features:
- Excel template generation with sample data
- Comprehensive validation (format, data types, business rules)
- Transaction-safe imports (rollback on ANY error)
- UPDATE mode for existing records
- Detailed error reporting with row numbers
- Audit logging for all imports
"""
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from decimal import Decimal

from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.core.models.products import Partner, PartnerType, Product, ProductType, UOM, Category
from app.core.models.bom import BOMHeader, BOMDetail, BOMType
from app.shared.audit import AuditLogger


class MasterdataImportService:
    """Service for masterdata bulk imports."""

    def __init__(self, db: Session, user_id: int):
        self.db = db
        self.user_id = user_id
        self.audit_logger = AuditLogger(db)

    # ====================== TEMPLATE GENERATION ======================

    def generate_suppliers_template(self) -> BytesIO:
        """Generate Excel template for suppliers import with sample data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suppliers"

        # Headers
        headers = [
            "supplier_code",
            "supplier_name",
            "supplier_type",
            "contact_person",
            "phone",
            "email",
            "address",
            "notes"
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sample data
        sample_data = [
            ["SUPP001", "PT Kain Jaya", "SUPPLIER", "John Doe", "081234567890", "john@kainjaya.com", "Jakarta Utara", "Fabric supplier"],
            ["SUPP002", "CV Label Indo", "SUPPLIER", "Jane Smith", "081298765432", "jane@labelindo.com", "Jakarta Barat", "Label specialist"],
            ["SUBB001", "UD Bordir Asri", "SUBCON", "Ahmad", "081211112222", "ahmad@bordirasri.com", "Tangerang", "Embroidery subcontractor"],
        ]
        for row in sample_data:
            ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_materials_template(self) -> BytesIO:
        """Generate Excel template for materials import with sample data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materials"

        # Headers
        headers = [
            "material_code",
            "material_name",
            "material_type",
            "uom",
            "category",
            "minimum_stock",
            "notes"
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sample data
        sample_data = [
            ["IKHR504", "KOHAIR 7MM RECYCLE D.BROWN", "RAW_MATERIAL", "YARD", "Fabric", "200", "Main fabric for bear"],
            ["THR001", "Thread Polyester Brown", "BAHAN_PENOLONG", "CONE", "Thread", "50", "Sewing thread"],
            ["FILL001", "Filling Dacron", "BAHAN_PENOLONG", "KG", "Filling", "20", "Stuffing material"],
        ]
        for row in sample_data:
            ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add instructions
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ["Field", "Description", "Valid Values", "Required"],
            ["material_code", "Unique material code", "Max 50 chars, alphanumeric", "YES"],
            ["material_name", "Material name", "Max 255 chars", "YES"],
            ["material_type", "Type of material", "RAW_MATERIAL, BAHAN_PENOLONG, WIP, FINISHED_GOODS", "YES"],
            ["uom", "Unit of measure", "PCS, YARD, METER, KG, GRAM, CONE, ROLL, BOX, CARTON", "YES"],
            ["category", "Material category", "Must match existing category name", "YES"],
            ["minimum_stock", "Minimum stock level", "Positive number", "NO (default: 0)"],
            ["notes", "Additional notes", "Text", "NO"],
        ]
        for row in instructions:
            ws2.append(row)

        # Style instructions
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(bold=True)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_articles_template(self) -> BytesIO:
        """Generate Excel template for articles import with sample data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"

        # Headers
        headers = [
            "article_code",
            "article_name",
            "description",
            "buyer",
            "category",
            "standard_packing",
            "parent_article_code",
            "notes"
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sample data
        sample_data = [
            ["40551542", "AFTONSPARV soft toy w astronaut suit 28 bear", "IKEA Bear with suit", "IKEA", "Soft Toys", "60", "", "Main article"],
            ["40551542-BODY", "AFTONSPARV Bear Body (WIP)", "Bear body without suit", "IKEA", "WIP", "60", "40551542", "Child of 40551542"],
            ["40551542-SUIT", "AFTONSPARV Astronaut Suit (WIP)", "Suit only", "IKEA", "WIP", "60", "40551542", "Child of 40551542"],
        ]
        for row in sample_data:
            ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_bom_template(self) -> BytesIO:
        """Generate Excel template for BOM import with sample data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"

        # Headers
        headers = [
            "article_code",
            "component_code",
            "quantity_required",
            "wastage_percent",
            "notes"
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sample data
        sample_data = [
            ["40551542", "IKHR504", "0.1466", "5", "Main fabric for body"],
            ["40551542", "THR001", "0.02", "0", "Sewing thread"],
            ["40551542", "FILL001", "0.05", "3", "Filling material"],
        ]
        for row in sample_data:
            ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ====================== VALIDATION ======================

    def validate_suppliers_data(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate suppliers dataframe.
        
        Returns:
            (is_valid, list_of_errors)
        """
        errors = []
        required_cols = ["supplier_code", "supplier_name", "supplier_type"]
        
        # Check required columns exist
        missing_cols = set(required_cols) - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return False, errors

        # Validate each row
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row (header is row 1)

            # Check supplier_code not empty
            if pd.isna(row['supplier_code']) or str(row['supplier_code']).strip() == '':
                errors.append(f"Row {row_num}: supplier_code is required")

            # Check supplier_name not empty
            if pd.isna(row['supplier_name']) or str(row['supplier_name']).strip() == '':
                errors.append(f"Row {row_num}: supplier_name is required")

            # Check supplier_type valid
            valid_types = ['SUPPLIER', 'SUBCON', 'CUSTOMER']
            if pd.isna(row['supplier_type']) or str(row['supplier_type']).upper() not in valid_types:
                errors.append(f"Row {row_num}: supplier_type must be one of {valid_types}")

            # Validate phone format (optional but if provided must be valid)
            if not pd.isna(row.get('phone')):
                phone = str(row['phone']).strip()
                if phone and not phone.replace('+', '').replace('-', '').replace(' ', '').isdigit():
                    errors.append(f"Row {row_num}: phone must contain only digits, +, -, and spaces")

        return len(errors) == 0, errors

    def validate_materials_data(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate materials dataframe."""
        errors = []
        required_cols = ["material_code", "material_name", "material_type", "uom", "category"]
        
        # Check required columns
        missing_cols = set(required_cols) - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return False, errors

        # Valid enums
        valid_types = ['RAW_MATERIAL', 'BAHAN_PENOLONG', 'WIP', 'FINISHED_GOODS']
        valid_uoms = ['PCS', 'YARD', 'METER', 'KG', 'GRAM', 'CONE', 'ROLL', 'BOX', 'CARTON', 'SET']

        # Get valid categories from database
        valid_categories = {cat.name for cat in self.db.query(Category).all()}

        for idx, row in df.iterrows():
            row_num = idx + 2

            # Check material_code
            if pd.isna(row['material_code']) or str(row['material_code']).strip() == '':
                errors.append(f"Row {row_num}: material_code is required")

            # Check material_name
            if pd.isna(row['material_name']) or str(row['material_name']).strip() == '':
                errors.append(f"Row {row_num}: material_name is required")

            # Check material_type
            if pd.isna(row['material_type']) or str(row['material_type']).upper() not in valid_types:
                errors.append(f"Row {row_num}: material_type must be one of {valid_types}")

            # Check uom
            if pd.isna(row['uom']) or str(row['uom']).upper() not in valid_uoms:
                errors.append(f"Row {row_num}: uom must be one of {valid_uoms}")

            # Check category exists
            if pd.isna(row['category']):
                errors.append(f"Row {row_num}: category is required")
            elif str(row['category']).strip() not in valid_categories:
                errors.append(f"Row {row_num}: category '{row['category']}' not found in database. Valid: {valid_categories}")

            # Check minimum_stock is positive number (if provided)
            if not pd.isna(row.get('minimum_stock')):
                try:
                    min_stock = float(row['minimum_stock'])
                    if min_stock < 0:
                        errors.append(f"Row {row_num}: minimum_stock must be >= 0")
                except ValueError:
                    errors.append(f"Row {row_num}: minimum_stock must be a number")

        return len(errors) == 0, errors

    def validate_bom_data(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate BOM dataframe."""
        errors = []
        required_cols = ["article_code", "component_code", "quantity_required"]
        
        # Check required columns
        missing_cols = set(required_cols) - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return False, errors

        # Get valid product codes
        valid_products = {p.code for p in self.db.query(Product).all()}

        for idx, row in df.iterrows():
            row_num = idx + 2

            # Check article_code exists
            if pd.isna(row['article_code']) or str(row['article_code']).strip() == '':
                errors.append(f"Row {row_num}: article_code is required")
            elif str(row['article_code']).strip() not in valid_products:
                errors.append(f"Row {row_num}: article_code '{row['article_code']}' not found in products")

            # Check component_code exists
            if pd.isna(row['component_code']) or str(row['component_code']).strip() == '':
                errors.append(f"Row {row_num}: component_code is required")
            elif str(row['component_code']).strip() not in valid_products:
                errors.append(f"Row {row_num}: component_code '{row['component_code']}' not found in products")

            # Check quantity_required
            if pd.isna(row['quantity_required']):
                errors.append(f"Row {row_num}: quantity_required is required")
            else:
                try:
                    qty = float(row['quantity_required'])
                    if qty <= 0:
                        errors.append(f"Row {row_num}: quantity_required must be > 0")
                except ValueError:
                    errors.append(f"Row {row_num}: quantity_required must be a number")

            # Check wastage_percent (optional, but if provided must be valid)
            if not pd.isna(row.get('wastage_percent')):
                try:
                    wastage = float(row['wastage_percent'])
                    if wastage < 0 or wastage > 100:
                        errors.append(f"Row {row_num}: wastage_percent must be between 0-100")
                except ValueError:
                    errors.append(f"Row {row_num}: wastage_percent must be a number")

        return len(errors) == 0, errors

    # ====================== IMPORT EXECUTION ======================

    def import_suppliers(self, file_content: bytes) -> Dict:
        """Import suppliers from Excel file.
        
        Returns:
            {
                "success": bool,
                "imported_count": int,
                "updated_count": int,
                "errors": List[str],
                "execution_time_ms": int
            }
        """
        start_time = datetime.now()

        try:
            # Read Excel
            df = pd.read_excel(BytesIO(file_content))
            
            # Validate
            is_valid, errors = self.validate_suppliers_data(df)
            if not is_valid:
                return {
                    "success": False,
                    "imported_count": 0,
                    "updated_count": 0,
                    "errors": errors,
                    "execution_time_ms": 0
                }

            imported_count = 0
            updated_count = 0
            errors = []

            # Begin transaction
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    supplier_code = str(row['supplier_code']).strip()
                    
                    # Check if exists
                    existing = self.db.query(Partner).filter(Partner.name == supplier_code).first()
                    
                    if existing:
                        # UPDATE mode
                        existing.name = str(row['supplier_name']).strip()
                        existing.type = PartnerType[str(row['supplier_type']).upper()]
                        existing.contact_person = str(row.get('contact_person', '')).strip() or None
                        existing.phone = str(row.get('phone', '')).strip() or None
                        existing.email = str(row.get('email', '')).strip() or None
                        existing.address = str(row.get('address', '')).strip() or None
                        updated_count += 1
                    else:
                        # INSERT mode
                        partner = Partner(
                            name=str(row['supplier_name']).strip(),
                            type=PartnerType[str(row['supplier_type']).upper()],
                            contact_person=str(row.get('contact_person', '')).strip() or None,
                            phone=str(row.get('phone', '')).strip() or None,
                            email=str(row.get('email', '')).strip() or None,
                            address=str(row.get('address', '')).strip() or None
                        )
                        self.db.add(partner)
                        imported_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    self.db.rollback()
                    return {
                        "success": False,
                        "imported_count": 0,
                        "updated_count": 0,
                        "errors": errors,
                        "execution_time_ms": 0
                    }

            # Commit transaction
            self.db.commit()

            # Audit log
            self.audit_logger.log_create(
                user_id=self.user_id,
                module="Masterdata Import",
                action="Import Suppliers",
                entity_type="Partner",
                entity_id=None,
                changes=f"Imported {imported_count} suppliers, Updated {updated_count} suppliers"
            )

            execution_time = (datetime.now() - start_time).total_seconds() * 1000

            return {
                "success": True,
                "imported_count": imported_count,
                "updated_count": updated_count,
                "errors": [],
                "execution_time_ms": int(execution_time)
            }

        except Exception as e:
            self.db.rollback()
            return {
                "success": False,
                "imported_count": 0,
                "updated_count": 0,
                "errors": [f"Fatal error: {str(e)}"],
                "execution_time_ms": 0
            }

    def import_materials(self, file_content: bytes) -> Dict:
        """Import materials from Excel file."""
        start_time = datetime.now()

        try:
            df = pd.read_excel(BytesIO(file_content))
            
            # Validate
            is_valid, errors = self.validate_materials_data(df)
            if not is_valid:
                return {
                    "success": False,
                    "imported_count": 0,
                    "updated_count": 0,
                    "errors": errors,
                    "execution_time_ms": 0
                }

            imported_count = 0
            updated_count = 0
            errors = []

            # Get category mapping
            categories = {cat.name: cat.id for cat in self.db.query(Category).all()}

            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    material_code = str(row['material_code']).strip()
                    
                    # Check if exists
                    existing = self.db.query(Product).filter(Product.code == material_code).first()
                    
                    category_id = categories.get(str(row['category']).strip())
                    if not category_id:
                        errors.append(f"Row {row_num}: Category '{row['category']}' not found")
                        continue

                    min_stock = float(row.get('minimum_stock', 0)) if not pd.isna(row.get('minimum_stock')) else 0

                    if existing:
                        # UPDATE mode
                        existing.name = str(row['material_name']).strip()
                        existing.type = ProductType[str(row['material_type']).upper()]
                        existing.uom = UOM[str(row['uom']).upper()]
                        existing.category_id = category_id
                        existing.min_stock = Decimal(str(min_stock))
                        updated_count += 1
                    else:
                        # INSERT mode
                        product = Product(
                            code=material_code,
                            name=str(row['material_name']).strip(),
                            type=ProductType[str(row['material_type']).upper()],
                            uom=UOM[str(row['uom']).upper()],
                            category_id=category_id,
                            min_stock=Decimal(str(min_stock)),
                            is_active=True
                        )
                        self.db.add(product)
                        imported_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    self.db.rollback()
                    return {
                        "success": False,
                        "imported_count": 0,
                        "updated_count": 0,
                        "errors": errors,
                        "execution_time_ms": 0
                    }

            self.db.commit()

            self.audit_logger.log_create(
                user_id=self.user_id,
                module="Masterdata Import",
                action="Import Materials",
                entity_type="Product",
                entity_id=None,
                changes=f"Imported {imported_count} materials, Updated {updated_count} materials"
            )

            execution_time = (datetime.now() - start_time).total_seconds() * 1000

            return {
                "success": True,
                "imported_count": imported_count,
                "updated_count": updated_count,
                "errors": [],
                "execution_time_ms": int(execution_time)
            }

        except Exception as e:
            self.db.rollback()
            return {
                "success": False,
                "imported_count": 0,
                "updated_count": 0,
                "errors": [f"Fatal error: {str(e)}"],
                "execution_time_ms": 0
            }

    def import_bom(self, file_content: bytes) -> Dict:
        """Import BOM from Excel file."""
        start_time = datetime.now()

        try:
            df = pd.read_excel(BytesIO(file_content))
            
            # Validate
            is_valid, errors = self.validate_bom_data(df)
            if not is_valid:
                return {
                    "success": False,
                    "imported_count": 0,
                    "updated_count": 0,
                    "errors": errors,
                    "execution_time_ms": 0
                }

            imported_count = 0
            updated_count = 0
            errors = []

            # Get product mapping
            products = {p.code: p.id for p in self.db.query(Product).all()}

            # Group by article_code (BOM Header)
            for article_code, group in df.groupby('article_code'):
                try:
                    article_id = products.get(str(article_code).strip())
                    if not article_id:
                        errors.append(f"Article '{article_code}' not found")
                        continue

                    # Check if BOM header exists
                    bom_header = self.db.query(BOMHeader).filter(
                        BOMHeader.product_id == article_id,
                        BOMHeader.is_active == True
                    ).first()

                    if not bom_header:
                        # Create new BOM header
                        bom_header = BOMHeader(
                            product_id=article_id,
                            bom_type=BOMType.MANUFACTURING,
                            qty_output=Decimal('1.0'),
                            is_active=True,
                            revision="Rev 1.0",
                            revised_by=self.user_id
                        )
                        self.db.add(bom_header)
                        self.db.flush()  # Get ID

                    # Process BOM details
                    for idx, row in group.iterrows():
                        component_id = products.get(str(row['component_code']).strip())
                        if not component_id:
                            errors.append(f"Component '{row['component_code']}' not found")
                            continue

                        qty_needed = Decimal(str(row['quantity_required']))
                        wastage = Decimal(str(row.get('wastage_percent', 0))) if not pd.isna(row.get('wastage_percent')) else Decimal('0')

                        # Check if detail exists
                        existing_detail = self.db.query(BOMDetail).filter(
                            BOMDetail.bom_header_id == bom_header.id,
                            BOMDetail.component_id == component_id
                        ).first()

                        if existing_detail:
                            # UPDATE
                            existing_detail.qty_needed = qty_needed
                            existing_detail.wastage_percent = wastage
                            updated_count += 1
                        else:
                            # INSERT
                            detail = BOMDetail(
                                bom_header_id=bom_header.id,
                                component_id=component_id,
                                qty_needed=qty_needed,
                                wastage_percent=wastage,
                                has_variants=False
                            )
                            self.db.add(detail)
                            imported_count += 1

                except Exception as e:
                    errors.append(f"Article '{article_code}': {str(e)}")
                    self.db.rollback()
                    return {
                        "success": False,
                        "imported_count": 0,
                        "updated_count": 0,
                        "errors": errors,
                        "execution_time_ms": 0
                    }

            self.db.commit()

            self.audit_logger.log_create(
                user_id=self.user_id,
                module="Masterdata Import",
                action="Import BOM",
                entity_type="BOMDetail",
                entity_id=None,
                changes=f"Imported {imported_count} BOM lines, Updated {updated_count} BOM lines"
            )

            execution_time = (datetime.now() - start_time).total_seconds() * 1000

            return {
                "success": True,
                "imported_count": imported_count,
                "updated_count": updated_count,
                "errors": [],
                "execution_time_ms": int(execution_time)
            }

        except Exception as e:
            self.db.rollback()
            return {
                "success": False,
                "imported_count": 0,
                "updated_count": 0,
                "errors": [f"Fatal error: {str(e)}"],
                "execution_time_ms": 0
            }
