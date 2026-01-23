"""Import/Export Module - CSV/Excel Import Export functionality
Supports: Products, BOM, Master Data, Inventory, Users
"""
import csv
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import require_permission
from app.core.models.bom import BOMDetail, BOMHeader
from app.core.models.products import Category, Product
from app.core.models.users import User
from app.core.models.warehouse import Location, StockQuant
from app.shared.audit import AuditLogger

router = APIRouter(prefix="/import-export", tags=["Import/Export"])

# ====================== IMPORT ENDPOINTS ======================

@router.post("/import/products", dependencies=[Depends(require_permission("import_export.import_data"))])
async def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.import_data"))
):
    """Import products from CSV/Excel file

    CSV Format:
    code,name,type,uom,category_id,min_stock
    BLU-SHARK,Blue Shark Plush,Finish Good,Pcs,1,100
    FAB-VEL-WHT,White Velvet Fabric,Raw Material,Meter,2,500

    Excel Format: Same columns as CSV
    """
    if file.filename.endswith('.csv'):
        return await _import_products_csv(file, db, current_user)
    elif file.filename.endswith(('.xlsx', '.xls')):
        return await _import_products_excel(file, db, current_user)
    else:
        raise HTTPException(status_code=400, detail="File must be CSV or Excel (.xlsx, .xls)")


async def _import_products_csv(file: UploadFile, db: Session, current_user: User):
    """Import products from CSV"""
    content = await file.read()
    decoded = content.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded))

    imported_count = 0
    errors = []
    audit_logger = AuditLogger(db)

    for row_num, row in enumerate(csv_reader, start=2):
        try:
            # Check if product exists
            existing = db.query(Product).filter(Product.code == row['code']).first()
            if existing:
                errors.append(f"Row {row_num}: Product {row['code']} already exists")
                continue

            # Validate category exists
            category = db.query(Category).filter(Category.id == int(row['category_id'])).first()
            if not category:
                errors.append(f"Row {row_num}: Category ID {row['category_id']} not found")
                continue

            # Create product
            product = Product(
                code=row['code'],
                name=row['name'],
                type=row['type'],
                uom=row['uom'],
                category_id=int(row['category_id']),
                min_stock=float(row.get('min_stock', 0))
            )
            db.add(product)
            imported_count += 1

            # Audit log
            audit_logger.log_create(
                user_id=current_user.id,
                module="Products",
                action="Import",
                entity_type="Product",
                entity_id=None,
                changes=f"Imported product {row['code']}"
            )

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return {
        "status": "success",
        "imported": imported_count,
        "errors": errors,
        "total_rows": row_num - 1
    }


async def _import_products_excel(file: UploadFile, db: Session, current_user: User):
    """Import products from Excel"""
    content = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active

    imported_count = 0
    errors = []
    audit_logger = AuditLogger(db)

    # Assume first row is header
    headers = [cell.value for cell in sheet[1]]

    for row_num in range(2, sheet.max_row + 1):
        try:
            row_data = {headers[i]: sheet.cell(row_num, i+1).value for i in range(len(headers))}

            # Check if product exists
            existing = db.query(Product).filter(Product.code == row_data['code']).first()
            if existing:
                errors.append(f"Row {row_num}: Product {row_data['code']} already exists")
                continue

            # Validate category
            category = db.query(Category).filter(Category.id == int(row_data['category_id'])).first()
            if not category:
                errors.append(f"Row {row_num}: Category ID {row_data['category_id']} not found")
                continue

            # Create product
            product = Product(
                code=row_data['code'],
                name=row_data['name'],
                type=row_data['type'],
                uom=row_data['uom'],
                category_id=int(row_data['category_id']),
                min_stock=float(row_data.get('min_stock', 0))
            )
            db.add(product)
            imported_count += 1

            # Audit log
            audit_logger.log_create(
                user_id=current_user.id,
                module="Products",
                action="Import",
                entity_type="Product",
                entity_id=None,
                changes=f"Imported product {row_data['code']}"
            )

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return {
        "status": "success",
        "imported": imported_count,
        "errors": errors,
        "total_rows": sheet.max_row - 1
    }


@router.post("/import/bom", dependencies=[Depends(require_permission("import_export.import_data"))])
async def import_bom(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.import_data"))
):
    """Import BOM (Bill of Materials) from CSV/Excel

    CSV Format:
    product_code,component_code,qty_needed,wastage_percent
    WIP-SEW-SHARK,FAB-VEL-WHT,2.5,5
    WIP-SEW-SHARK,THR-BLU-001,0.15,2

    Excel Format: Same columns as CSV
    """
    if file.filename.endswith('.csv'):
        return await _import_bom_csv(file, db, current_user)
    elif file.filename.endswith(('.xlsx', '.xls')):
        return await _import_bom_excel(file, db, current_user)
    else:
        raise HTTPException(status_code=400, detail="File must be CSV or Excel (.xlsx, .xls)")


async def _import_bom_csv(file: UploadFile, db: Session, current_user: User):
    """Import BOM from CSV"""
    content = await file.read()
    decoded = content.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded))

    imported_count = 0
    errors = []
    audit_logger = AuditLogger(db)

    for row_num, row in enumerate(csv_reader, start=2):
        try:
            # Find product
            product = db.query(Product).filter(Product.code == row['product_code']).first()
            if not product:
                errors.append(f"Row {row_num}: Product {row['product_code']} not found")
                continue

            # Find component
            component = db.query(Product).filter(Product.code == row['component_code']).first()
            if not component:
                errors.append(f"Row {row_num}: Component {row['component_code']} not found")
                continue

            # Check if BOM header exists
            bom_header = db.query(BOMHeader).filter(BOMHeader.product_id == product.id).first()
            if not bom_header:
                # Create BOM header
                bom_header = BOMHeader(
                    product_id=product.id,
                    bom_type="Manufacturing",
                    qty_output=1.0,
                    is_active=True,
                    revision="1.0"
                )
                db.add(bom_header)
                db.flush()

            # Create BOM detail
            bom_detail = BOMDetail(
                bom_header_id=bom_header.id,
                component_id=component.id,
                qty_needed=float(row['qty_needed']),
                wastage_percent=float(row.get('wastage_percent', 0))
            )
            db.add(bom_detail)
            imported_count += 1

            # Audit log
            audit_logger.log_create(
                user_id=current_user.id,
                module="BOM",
                action="Import",
                entity_type="BOMDetail",
                entity_id=None,
                changes=f"Imported BOM for {row['product_code']}"
            )

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return {
        "status": "success",
        "imported": imported_count,
        "errors": errors,
        "total_rows": row_num - 1
    }


async def _import_bom_excel(file: UploadFile, db: Session, current_user: User):
    """Import BOM from Excel"""
    content = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active

    imported_count = 0
    errors = []
    audit_logger = AuditLogger(db)

    headers = [cell.value for cell in sheet[1]]

    for row_num in range(2, sheet.max_row + 1):
        try:
            row_data = {headers[i]: sheet.cell(row_num, i+1).value for i in range(len(headers))}

            # Find product
            product = db.query(Product).filter(Product.code == row_data['product_code']).first()
            if not product:
                errors.append(f"Row {row_num}: Product {row_data['product_code']} not found")
                continue

            # Find component
            component = db.query(Product).filter(Product.code == row_data['component_code']).first()
            if not component:
                errors.append(f"Row {row_num}: Component {row_data['component_code']} not found")
                continue

            # Check if BOM header exists
            bom_header = db.query(BOMHeader).filter(BOMHeader.product_id == product.id).first()
            if not bom_header:
                bom_header = BOMHeader(
                    product_id=product.id,
                    bom_type="Manufacturing",
                    qty_output=1.0,
                    is_active=True,
                    revision="1.0"
                )
                db.add(bom_header)
                db.flush()

            # Create BOM detail
            bom_detail = BOMDetail(
                bom_header_id=bom_header.id,
                component_id=component.id,
                qty_needed=float(row_data['qty_needed']),
                wastage_percent=float(row_data.get('wastage_percent', 0))
            )
            db.add(bom_detail)
            imported_count += 1

            # Audit log
            audit_logger.log_create(
                user_id=current_user.id,
                module="BOM",
                action="Import",
                entity_type="BOMDetail",
                entity_id=None,
                changes=f"Imported BOM for {row_data['product_code']}"
            )

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()

    return {
        "status": "success",
        "imported": imported_count,
        "errors": errors,
        "total_rows": sheet.max_row - 1
    }


# ====================== EXPORT ENDPOINTS ======================

@router.get("/export/products")
async def export_products(
    format: str = Query("csv", regex="^(csv|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.export_data"))
):
    """Export all products to CSV or Excel

    Query Parameters:
    - format: csv or excel
    """
    products = db.query(Product).all()

    if format == "csv":
        return _export_products_csv(products)
    else:
        return _export_products_excel(products)


def _export_products_csv(products: list[Product]):
    """Export products to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['code', 'name', 'type', 'uom', 'category_id', 'min_stock', 'created_at'])

    # Data
    for product in products:
        writer.writerow([
            product.code,
            product.name,
            product.type,
            product.uom,
            product.category_id,
            product.min_stock,
            product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else ''
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def _export_products_excel(products: list[Product]):
    """Export products to Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"

    # Header
    headers = ['code', 'name', 'type', 'uom', 'category_id', 'min_stock', 'created_at']
    sheet.append(headers)

    # Data
    for product in products:
        sheet.append([
            product.code,
            product.name,
            product.type,
            product.uom,
            product.category_id,
            product.min_stock,
            product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else ''
        ])

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )


@router.get("/export/bom")
async def export_bom(
    format: str = Query("csv", regex="^(csv|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.export_data"))
):
    """Export BOM (Bill of Materials) to CSV or Excel

    Query Parameters:
    - format: csv or excel
    """
    # Query BOM with product and component info
    bom_data = db.query(
        BOMHeader, BOMDetail, Product,
        Product.code.label('component_code')
    ).join(
        BOMDetail, BOMHeader.id == BOMDetail.bom_header_id
    ).join(
        Product, BOMHeader.product_id == Product.id
    ).join(
        Product, BOMDetail.component_id == Product.id, isouter=True
    ).all()

    if format == "csv":
        return _export_bom_csv(bom_data, db)
    else:
        return _export_bom_excel(bom_data, db)


def _export_bom_csv(bom_data, db: Session):
    """Export BOM to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['product_code', 'product_name', 'component_code', 'component_name', 'qty_needed', 'wastage_percent'])

    # Data
    for bom_header, bom_detail, product, component_code in bom_data:
        component = db.query(Product).filter(Product.id == bom_detail.component_id).first()
        writer.writerow([
            product.code,
            product.name,
            component.code if component else '',
            component.name if component else '',
            bom_detail.qty_needed,
            bom_detail.wastage_percent
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bom_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def _export_bom_excel(bom_data, db: Session):
    """Export BOM to Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "BOM"

    # Header
    headers = ['product_code', 'product_name', 'component_code', 'component_name', 'qty_needed', 'wastage_percent']
    sheet.append(headers)

    # Data
    for bom_header, bom_detail, product, component_code in bom_data:
        component = db.query(Product).filter(Product.id == bom_detail.component_id).first()
        sheet.append([
            product.code,
            product.name,
            component.code if component else '',
            component.name if component else '',
            bom_detail.qty_needed,
            bom_detail.wastage_percent
        ])

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bom_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )


@router.get("/export/inventory")
async def export_inventory(
    format: str = Query("csv", regex="^(csv|excel)$"),
    location_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.export_data"))
):
    """Export current inventory to CSV or Excel

    Query Parameters:
    - format: csv or excel
    - location_id: Optional filter by location
    """
    query = db.query(StockQuant, Product, Location).join(
        Product, StockQuant.product_id == Product.id
    ).join(
        Location, StockQuant.location_id == Location.id
    )

    if location_id:
        query = query.filter(StockQuant.location_id == location_id)

    inventory_data = query.all()

    if format == "csv":
        return _export_inventory_csv(inventory_data)
    else:
        return _export_inventory_excel(inventory_data)


def _export_inventory_csv(inventory_data):
    """Export inventory to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['product_code', 'product_name', 'location', 'qty_available', 'qty_reserved', 'uom'])

    # Data
    for stock_quant, product, location in inventory_data:
        writer.writerow([
            product.code,
            product.name,
            location.name,
            stock_quant.qty_available,
            stock_quant.qty_reserved,
            product.uom
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def _export_inventory_excel(inventory_data):
    """Export inventory to Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"

    # Header
    headers = ['product_code', 'product_name', 'location', 'qty_available', 'qty_reserved', 'uom']
    sheet.append(headers)

    # Data
    for stock_quant, product, location in inventory_data:
        sheet.append([
            product.code,
            product.name,
            location.name,
            stock_quant.qty_available,
            stock_quant.qty_reserved,
            product.uom
        ])

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )


@router.get("/export/users")
async def export_users(
    format: str = Query("csv", regex="^(csv|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("import_export.export_data"))
):
    """Export users to CSV or Excel (for backup/audit)

    Query Parameters:
    - format: csv or excel
    """
    users = db.query(User).all()

    if format == "csv":
        return _export_users_csv(users)
    else:
        return _export_users_excel(users)


def _export_users_csv(users: list[User]):
    """Export users to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['username', 'email', 'full_name', 'role', 'department', 'is_active', 'created_at'])

    # Data
    for user in users:
        writer.writerow([
            user.username,
            user.email,
            user.full_name,
            user.role,
            user.department,
            user.is_active,
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def _export_users_excel(users: list[User]):
    """Export users to Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    # Header
    headers = ['username', 'email', 'full_name', 'role', 'department', 'is_active', 'created_at']
    sheet.append(headers)

    # Data
    for user in users:
        sheet.append([
            user.username,
            user.email,
            user.full_name,
            user.role,
            user.department,
            user.is_active,
            user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
        ])

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )
